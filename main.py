import time
import datetime
import os
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver import ActionChains
import requests
from bs4 import BeautifulSoup
import lxml
import csv
import openpyxl
from openpyxl import load_workbook


url_headers = {
    "Accept": "*/*",
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.0.0 Safari/537.36"
}

links = []
s = Service(executable_path=r'C:\Users\1\Testing_learning\venv\Parcing_yandexMaps\chromedriver.exe')
driver = webdriver.Chrome(service=s)

def get_source():

    # s = Service(executable_path=r'C:\Users\1\Testing_learning\venv\Parcing_yandexMaps\chromedriver.exe')
    # driver = webdriver.Chrome(service=s)

    try:
        driver.maximize_window()
        driver.get("https://yandex.kz/maps/")
        time.sleep(3)
        input = driver.find_element(By.XPATH, "/html[1]/body[1]/div[1]/div[2]/div[2]/header[1]/div[1]/div[1]/div[1]/form[1]/div[1]/div[1]/span[1]/span[1]/input[1]")
        time.sleep(2)
        input.send_keys("Актау бар")
        time.sleep(2)
        input.send_keys(Keys.ENTER)
        time.sleep(3)

        while True:
                # в бесконечном цикле проходим до конца страницы
                div_element_ol = driver.find_elements(By.CLASS_NAME, 'seo-pagination-view')# прогружает новую страницу с данными
                print(f'количество вложенных списков - {len(div_element_ol)}')
                divs_element_placeholder = driver.find_elements(By.CLASS_NAME, 'search-snippet-view__placeholder')
                print(f'количество неотрытых карточек - {len(divs_element_placeholder)}') # Раскрывает списки карточек из прогруженных страниц
                for index in range(0, len(divs_element_placeholder), 2):
                    actions = ActionChains(driver)
                    driver.implicitly_wait(30)
                    actions.move_to_element(divs_element_placeholder[index]).perform()
                    time.sleep(1)
                print(len(divs_element_placeholder))
                trigger = driver.find_elements(By.CLASS_NAME, 'add-business-view__link')
                if trigger and (len(divs_element_placeholder) == 0):
                    with open("source_page.html", mode='w', encoding='utf-8') as file:
                        file.write(driver.page_source)
                    break
                else:
                    actions = ActionChains(driver)
                    actions.move_to_element(div_element_ol[0]).perform()
                    time.sleep(3)
    except Exception as exc:
        print(exc)
    finally:
        driver.close()
        driver.quit()

def collect_data():

    data = []
    data_value = []

    url_headers = {
        "Accept": "*/*",
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.0.0 Safari/537.36"
    }

    with open('source_page.html', encoding='utf-8') as file:
        src = file.read()

    soup = BeautifulSoup(src, "lxml")
    cards = soup.find_all('li', class_='search-snippet-view')
    for i in cards:
        link = "https://yandex.kz" + i.find('a').get('href')
        links.append(link)

    count = 0
    for i in links:

        url = ''.join(i).strip()
        driver.get(url)
        time.sleep(3)
        try:
            name = driver.find_element(By.CLASS_NAME, "orgpage-header-view__header").text
        except Exception:
            name = "Нет данных"
        time.sleep(1)
        try:
            address = driver.find_element(By.CLASS_NAME, "orgpage-header-view__address").text
        except Exception:
            address = "Нет данных"
        time.sleep(2)
        try:
            rating = driver.find_element(By.CLASS_NAME, "business-rating-badge-view").text.replace("Рейтинг","").strip()
        except Exception:
            rating = "Нет данных"
        try:
            tel = driver.find_element(By.CLASS_NAME, "orgpage-phones-view").text.replace("Показать телефон","").strip()
        except Exception:
            tel = "Нет данных"
        try:
            schedule = driver.find_element(By.CLASS_NAME, "business-working-status-view").text.strip()
        except Exception:
            schedule = "Нет данных"

        extend_info = driver.find_element(By.LINK_TEXT,"Подробнее об организации")
        extend_info.send_keys(Keys.ENTER)
        time.sleep(3)
        url2 = driver.current_url

        response = requests.get(url=url2, headers=url_headers)
        soup2 = BeautifulSoup(response.text, "lxml")
        beer_price = 0
        prices = 0
        list_valued = soup2.find_all('div', class_="business-features-view__valued")


        if list_valued:
            for i in list_valued:
                if i.find(text='Цена бокала пива'):
                    beer_price = i.find(class_="business-features-view__valued-value").text
                if i.find(text='Средний счёт'):
                    prices = i.find(class_="business-features-view__valued-value").text


        data.append({
            "Наименование": name,
            "Адрес": address,
            "Рейтинг": rating,
            "Телефон": tel,
            "График работы": schedule,
            "Цена бокала пива": beer_price,
            "Цены": prices
         })
        count +=1
        print(f'обработано ссылок {count} из {len(links)}')


    book = openpyxl.Workbook()  # создание файла
    sheet_1 = book.create_sheet("Данные")  # создание вкладки
    book.remove(book.active)  # удаление пустой вкладки

    headers = list(data[0].keys())  # добавляем заголовки
    for col_num, header in enumerate(headers, 1):
        sheet_1.cell(row=1, column=col_num, value=header)

    for row_num, row_data in enumerate(data, 2):
        for col_num, cell_value in enumerate(row_data.values(), 1):
            sheet_1.cell(row=row_num, column=col_num, value=cell_value)
    book.save('data.xlsx')
    book.close()
    print("Файл записан, работа завершена")


def main():
    start_time = time.time()
    # get_source()
    collect_data()
    end_time = time.time()
    total_time = str(end_time - start_time)
    print(f'Время работы составило: {total_time}')


if __name__ =="__main__":
    main()